from openpyxl import Workbook

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws["A1"] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime

ws["A2"] = datetime.datetime.now()

ws = wb.create_sheet("test")

start_num = 1
for row in range(10):
    for col in range(10):
        value = start_num + (row * 10) + col
        ws.cell(row=row + 1, column=col + 1, value=value)  # ワークシートに値を設定

# 保存
wb.save("sample.xlsx")
